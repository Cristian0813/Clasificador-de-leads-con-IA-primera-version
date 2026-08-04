"""
Clasificador Automático de Correos/Leads con IA
-------------------------------------------------
Lee correos no leídos de una bandeja (IMAP), los clasifica por urgencia
y categoría usando la API de Anthropic, y guarda el resultado en un
Excel ordenado por prioridad. Ideal para negocios que reciben muchos
correos (inmobiliarias, e-commerce, clínicas, abogados, agencias) y
necesitan saber a quién responder primero.

Uso:
    python main.py

Requiere un archivo .env (ver .env.example) con las credenciales.
"""

import imaplib
import email
import os
import logging
from datetime import datetime
from email.header import decode_header

import openpyxl
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv
import anthropic

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler("clasificador.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

IMAP_SERVER = os.getenv("IMAP_SERVER", "imap.gmail.com")
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_APP_PASSWORD")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
MAX_CORREOS = int(os.getenv("MAX_CORREOS", "20"))
OUTPUT_FILE = "leads_clasificados.xlsx"

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

PROMPT_CLASIFICACION = """Eres un asistente que clasifica correos entrantes de un negocio.
Analiza el siguiente correo y responde SOLO con este formato exacto, sin texto adicional:

CATEGORIA: [Venta/Consulta/Queja/Soporte/Spam/Otro]
URGENCIA: [Alta/Media/Baja]
RESUMEN: [una frase de máximo 15 palabras resumiendo qué quiere el remitente]

Correo:
Asunto: {asunto}
De: {remitente}
Cuerpo: {cuerpo}
"""


def conectar_imap():
    """Conecta a la bandeja de entrada vía IMAP con reintentos."""
    intentos = 3
    for intento in range(1, intentos + 1):
        try:
            mail = imaplib.IMAP4_SSL(IMAP_SERVER)
            mail.login(EMAIL_USER, EMAIL_PASS)
            mail.select("inbox")
            logger.info("Conexión IMAP exitosa")
            return mail
        except imaplib.IMAP4.error as e:
            logger.warning(f"Intento {intento}/{intentos} fallido: {e}")
            if intento == intentos:
                raise
    return None


def decodificar_texto(texto):
    """Decodifica encabezados de correo que pueden venir en distintos charsets."""
    if texto is None:
        return ""
    partes = decode_header(texto)
    resultado = ""
    for contenido, codificacion in partes:
        if isinstance(contenido, bytes):
            resultado += contenido.decode(codificacion or "utf-8", errors="ignore")
        else:
            resultado += contenido
    return resultado


def extraer_cuerpo(msg):
    """Extrae el cuerpo de texto plano de un mensaje, sea simple o multipart."""
    if msg.is_multipart():
        for parte in msg.walk():
            tipo = parte.get_content_type()
            disposicion = str(parte.get("Content-Disposition"))
            if tipo == "text/plain" and "attachment" not in disposicion:
                try:
                    return parte.get_payload(decode=True).decode(errors="ignore")
                except Exception:
                    continue
        return ""
    try:
        return msg.get_payload(decode=True).decode(errors="ignore")
    except Exception:
        return ""


def clasificar_con_ia(asunto, remitente, cuerpo):
    """Envía el correo a la API de Anthropic y parsea la clasificación."""
    cuerpo_recortado = cuerpo[:1500]
    try:
        respuesta = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=200,
            messages=[{
                "role": "user",
                "content": PROMPT_CLASIFICACION.format(
                    asunto=asunto, remitente=remitente, cuerpo=cuerpo_recortado
                ),
            }],
        )
        texto = respuesta.content[0].text
        datos = {"categoria": "Otro", "urgencia": "Media", "resumen": ""}
        for linea in texto.splitlines():
            if linea.startswith("CATEGORIA:"):
                datos["categoria"] = linea.split(":", 1)[1].strip()
            elif linea.startswith("URGENCIA:"):
                datos["urgencia"] = linea.split(":", 1)[1].strip()
            elif linea.startswith("RESUMEN:"):
                datos["resumen"] = linea.split(":", 1)[1].strip()
        return datos
    except Exception as e:
        logger.error(f"Error clasificando correo '{asunto}': {e}")
        return {"categoria": "Error", "urgencia": "Media", "resumen": "No se pudo clasificar"}


def guardar_en_excel(resultados):
    """Guarda los resultados en un Excel, ordenados por urgencia."""
    orden_urgencia = {"Alta": 0, "Media": 1, "Baja": 2}
    resultados.sort(key=lambda r: orden_urgencia.get(r["urgencia"], 3))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    encabezados = ["Fecha", "Remitente", "Asunto", "Categoría", "Urgencia", "Resumen"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    colores_urgencia = {
        "Alta": PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid"),
        "Media": PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"),
        "Baja": PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"),
    }

    for r in resultados:
        ws.append([r["fecha"], r["remitente"], r["asunto"], r["categoria"], r["urgencia"], r["resumen"]])
        fila = ws.max_row
        color = colores_urgencia.get(r["urgencia"])
        if color:
            ws.cell(row=fila, column=5).fill = color

    for columna in ws.columns:
        ancho_max = max(len(str(c.value)) for c in columna if c.value) + 2
        ws.column_dimensions[columna[0].column_letter].width = min(ancho_max, 50)

    wb.save(OUTPUT_FILE)
    logger.info(f"Excel guardado: {OUTPUT_FILE}")


def main():
    logger.info("Iniciando clasificación de correos...")
    if not all([EMAIL_USER, EMAIL_PASS, ANTHROPIC_API_KEY]):
        logger.error("Faltan credenciales en el archivo .env. Revisa .env.example")
        return

    mail = conectar_imap()
    _, mensajes = mail.search(None, "UNSEEN")
    ids = mensajes[0].split()[-MAX_CORREOS:]
    logger.info(f"{len(ids)} correos no leídos encontrados (máximo {MAX_CORREOS})")

    resultados = []
    for num in ids:
        _, datos_msg = mail.fetch(num, "(RFC822)")
        msg = email.message_from_bytes(datos_msg[0][1])

        asunto = decodificar_texto(msg["Subject"])
        remitente = decodificar_texto(msg["From"])
        cuerpo = extraer_cuerpo(msg)
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

        clasificacion = clasificar_con_ia(asunto, remitente, cuerpo)
        resultados.append({
            "fecha": fecha,
            "remitente": remitente,
            "asunto": asunto,
            **clasificacion,
        })
        logger.info(f"Clasificado: {asunto[:40]} -> {clasificacion['urgencia']}")

    mail.logout()

    if resultados:
        guardar_en_excel(resultados)
        print(f"\n✅ Listo. {len(resultados)} correos clasificados en '{OUTPUT_FILE}'")
    else:
        print("\nNo hay correos nuevos para clasificar.")


if __name__ == "__main__":
    main()
